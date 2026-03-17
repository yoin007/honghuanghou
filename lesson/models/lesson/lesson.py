# _*_ coding: utf-8 _*_
# @Author: Tech
# @Time: 2025/10/18 16:20

import asyncio
import hashlib
import logging
import os
import re
import shutil
import sqlite3
import threading
import time
import warnings
from datetime import datetime, timedelta
import pandas as pd

logger = logging.getLogger(__name__)
from concurrent.futures import ThreadPoolExecutor
from functools import wraps
from openpyxl import load_workbook
from config.config import Config
from config.log import LogConfig
from client import down_file
from models.manage.member import Member, check_permission
from sendqueue import send_app_msg, send_text, send_image, send_file
from utils.fonts import get_cached_chinese_font

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

log = LogConfig().get_logger()
config = Config()

# send_text = print
# send_image = print


def error_handler(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        notify = kwargs.pop('notify', False)
        try:
            return func(*args, **kwargs)
        except Exception as e:
            error_msg = f"函数 {func.__name__} 执行出错: {e}"
            if notify and len(args) > 0 and hasattr(args[0], "notify_admins"):
                args[0].notify_admins(error_msg, log_level="error")
            else:
                log.error(error_msg)
            return None
    return wrapper

class Lesson:
    _instance = None
    _lock = threading.Lock()
    def __new__(cls, *args, **kwargs):
        with cls._lock:
            if not cls._instance:
                cls._instance = super().__new__(cls)
        return cls._instance

    def __init__(self):
        if hasattr(self, "_initialized"):
            return
        self._initialized = True
        self.lesson_dir = config.get_cross_platform_path("lesson_dir", "lesson.yaml")
        self.admin = config.get_config("lesson_admin", "lesson.yaml")
        self.week_change = config.get_config("week_change", "lesson.yaml")
        self.create_c_month_dir(notify=True)
        self.old_schedule = None  # 旧的课程表, 用于对比更新

        # 缓存数据
        self.cache_datas = {}

        self.refresh_cache()

    @error_handler
    def create_c_month_dir(self, month=""):
        """创建当前月份目录"""
        if month == "":
            now = datetime.now()
            c_month = now.strftime("%Y%m")
        else:
            c_month = month
        c_month_dir = os.path.join(self.lesson_dir, c_month)
        if not os.path.exists(c_month_dir):
            os.makedirs(c_month_dir, exist_ok=True)
            os.makedirs(os.path.join(c_month_dir, "class_schedule"), exist_ok=True)
            os.makedirs(os.path.join(c_month_dir, "schedule_history"), exist_ok=True)
            os.makedirs(os.path.join(c_month_dir, "temp"), exist_ok=True)
            self.notify_admins(f"创建目录成功: {c_month_dir}", log_level="info")
            return c_month_dir
        else:
            return ""
    
    def notify_admins(self, msg: str, log_level="info"):
        """通知管理员"""
        for admin in self.admin:
            send_text(msg, admin)
        if log_level == "info":
            log.info(msg)
        elif log_level == "error":
            log.error(msg)
        elif log_level == "warning":
            log.warning(msg)

    def load_excel_file(self, file_path, sheet_name=0, index_col=None):
        """加载Excel文件"""
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, index_col=index_col)
            return df
        except Exception as e:
            self.notify_admins(f"加载Excel文件失败: {e}")
            log.error(f"加载Excel文件失败: {e}")
            return pd.DataFrame()

    @property
    def now(self):
        """获取当前时间"""
        return datetime.now()
    
    @property
    def c_month(self):
        """获取当前月份"""
        return self.now.strftime("%Y%m")

    @property
    def week_info(self):
        """获取周次信息"""
        today = self.now.replace(hour=0, minute=0, second=0, microsecond=0)
        week_number = today.isocalendar()[1] + self.week_change
        monday = today - timedelta(days=today.weekday())
        monday_timestamp = int(monday.timestamp())
        sunday_timestamp = int((monday + timedelta(days=7)).timestamp()) -1
        current = [week_number, monday.strftime("%Y%m%d"), monday_timestamp, sunday_timestamp, 0] # 0 表示当前周
        next_week_number = week_number + 1
        next_monday = monday + timedelta(days=7)
        next_monday_timestamp = int(next_monday.timestamp())
        next_sunday_timestamp = int((next_monday + timedelta(days=7)).timestamp()) -1
        next = [next_week_number, next_monday.strftime("%Y%m%d"), next_monday_timestamp, next_sunday_timestamp, 1] # 1 表示下一周
        week_info_list = [current, next]
        return week_info_list
    
    @property
    def teacher_template(self):
        """获取教师模板"""
        file_path = os.path.join(self.lesson_dir, "checkTemplate.xlsx")
        return self.load_excel_file(file_path, sheet_name="teachers")
    
    @property
    def class_template(self):
        """获取班级模板"""
        file_path = os.path.join(self.lesson_dir, "checkTemplate.xlsx")
        return self.load_excel_file(file_path, sheet_name="class")
    
    @property
    def students(self):
        """获取学生列表"""
        file_path = os.path.join(self.lesson_dir, "checkTemplate.xlsx")
        stu_df = self.load_excel_file(file_path, sheet_name="students")
        if stu_df is None or stu_df.empty:
            return pd.DataFrame()
        
        # 转换 roomid 和 rpid 为整数再转字符串，去除小数点
        def safe_int_str(x):
            try:
                if pd.isna(x):
                    return ""
                if isinstance(x, (int, float)):
                    return str(int(x))
                return str(x)
            except:
                return str(x)

        for col in ['sid', 'rpid']:
            if col in stu_df.columns:
                stu_df[col] = stu_df[col].apply(safe_int_str)
        return stu_df
    

    @property
    def time_table(self):
        """获取时间表"""
        file_path = os.path.join(self.lesson_dir, "checkTemplate.xlsx")
        return self.load_excel_file(file_path, sheet_name="class_time")
    
    @property
    def members(self):
        """加载会员"""
        with Member() as m:
            m.update_contacts()
            members_list = m.member_info()
            if members_list is None:
                return pd.DataFrame()
            columns = m.member_columns()
        return pd.DataFrame(members_list, columns=columns)
    
    @property
    def schedule_file(self):
        """获取课程表文件"""
        c_s_file = ''
        n_s_file = ''
        c_month = self.get_cache_data("c_month")
        schedule_dir = os.path.join(self.lesson_dir, c_month, "class_schedule")
        week_info = self.get_cache_data("week_info")
        if not os.path.exists(schedule_dir):
            return [c_s_file, n_s_file]
        for f in os.listdir(schedule_dir):
            if week_info[0][1] in f:
                c_s_file = os.path.join(schedule_dir, f)
            if week_info[1][1] in f:
                n_s_file = os.path.join(schedule_dir, f)
        return [c_s_file, n_s_file]
    
    @property
    def replace_dict(self):
        """获取替换字典"""
        replace_template = self.load_excel_file(os.path.join(self.lesson_dir, "checkTemplate.xlsx"), sheet_name="replace")
        return dict(zip(replace_template["string"], replace_template["replace"]))
    
    @property
    def ignore_subjects(self):
        """获取忽略列表"""
        ignore_template = self.load_excel_file(os.path.join(self.lesson_dir, "checkTemplate.xlsx"), sheet_name="ignore")
        return ignore_template["subject"].tolist()
    
    def format_schedule(self, df_schedule:pd.DataFrame, week_flag:str, replace_flag:bool=True, ignore:bool=False):
        """格式化课程表（优化版）"""
        replace_dict = self.get_cache_data("replace_dict") if replace_flag else {}
        ignore_subjects = self.get_cache_data("ignore_subjects") if ignore else set()
        week_flag_str = f"({week_flag})"

        def process(x):
            # 非字符串直接返回
            if not isinstance(x, str):
                return x
            # 清理字符串
            result = x.strip().replace('（', '(').replace('）', ')').replace(' ', '')
            result = re.compile(r"[ \s\n\r\t]+" ).sub("", result)
            # 处理周标记
            subjects = result.split("/")
            if len(subjects) == 2:
                for subject in subjects:
                    if week_flag_str in subject:
                        return subject.replace(week_flag_str, "").strip()
            # 替换内容
            for k, v in replace_dict.items():
                result = result.replace(k, v)
            # 忽略学科
            if ignore and result in ignore_subjects:
                return "-"
            return result.strip()

        df_schedule = df_schedule.map(process)
        df_schedule = df_schedule.fillna("-")
        df_schedule.to_excel('temp.xlsx')
        return df_schedule
    
    def schedule_data(self, week_next:int=0, replace_flag:bool=True, ignore:bool=False):
        """获取当前周课程表文件"""
        schedule_file = self.get_cache_data("schedule_file")[week_next]
        if schedule_file == '':
            return pd.DataFrame()
        week_num = self.get_cache_data("week_info")[week_next][0]
        week_flag = "单" if week_num % 2 == 1 else "双"
        df = self.load_excel_file(schedule_file)
        df = self.format_schedule(df, week_flag, replace_flag, ignore)
        return df
    
    @property
    def current_schedule(self):
        """获取当前周课程表"""
        df = self.schedule_data(week_next=0)   
        return df
    
    @property
    def next_schedule(self):
        """获取下一周课程表"""
        df = self.schedule_data(week_next=1)   
        return df
    
    @property
    def today_schedule(self):
        """获取今日课程表"""
        df = self.get_cache_data("current_schedule")
        if df is None or df.empty:
            return pd.DataFrame()
        if "date" not in df.columns:
            return pd.DataFrame()
        today = self.get_cache_data("now").strftime("%d")
        df["date"] = df["date"].astype(int)
        today_df = df[df["date"] == int(today)]
        return today_df

    def get_weekday_schedule(self, target_date=None):
        """获取指定日期的课程表（支持传入日期或周几 1-7）"""
        df = self.get_cache_data("current_schedule")
        if df is None or df.empty:
            return pd.DataFrame()
        
        if target_date is None:
            now = self.get_cache_data("now")
            target_day = now.day
            target_weekday = now.weekday() + 1
        elif isinstance(target_date, int):
            target_weekday = target_date
            # 当 target_date 是整数（表示周几）时，不设置 target_day，后续只根据 weekday 筛选
            target_day = None
        else:
            try:
                if isinstance(target_date, str):
                    target_dt = datetime.strptime(target_date, "%Y%m%d")
                else:
                    target_dt = target_date
                target_day = target_dt.day
                target_weekday = target_dt.weekday() + 1
            except:
                now = self.get_cache_data("now")
                target_day = now.day
                target_weekday = now.weekday() + 1
        
        if "date" not in df.columns:
            return pd.DataFrame()
        
        df["date"] = df["date"].astype(int)
        
        if "weekday" in df.columns:
            df["weekday"] = df["weekday"].astype(int)
            if target_day is not None:
                schedule_df = df[(df["date"] == target_day) & (df["weekday"] == target_weekday)]
            else:
                # 当 target_day 为 None 时，只根据 weekday 筛选
                schedule_df = df[df["weekday"] == target_weekday]
        else:
            schedule_df = df[df["date"] == target_day]
        
        return schedule_df

    def refresh_cache(self):
        """刷新缓存"""
        self.cache_datas.clear()
        self.cache_datas = {
            "now": self.now,
            "c_month": self.c_month,
            "week_info": self.week_info,
            "members": self.members,
            "teacher_template": self.teacher_template,
            "class_template": self.class_template,
            "students": self.students,
            "replace_dict": self.replace_dict,
            "ignore_subjects": self.ignore_subjects,
            "time_table": self.time_table,
            "schedule_file": self.schedule_file,
            "current_schedule": self.current_schedule,
            "next_schedule": self.next_schedule,
            "today_schedule": self.today_schedule,
        }
        print(self.cache_datas["now"])
        log.info(f"缓存已刷新")
        return 1

    def get_cache_data(self,key):
        """获取缓存数据"""
        result = self.cache_datas.get(key, None)
        if result is None:
            if key == "now":
                return self.now
            if key == "c_month":
                return self.c_month
            if key == "week_info":
                return self.week_info
            if key == "teacher_template":
                return self.teacher_template
            if key == "class_template":
                return self.class_template
            if key == "students":
                return self.students
            if key == "time_table":
                return self.time_table
            if key == "members":
                return self.members
            if key == "schedule_file":
                return self.schedule_file
            if key == "current_schedule":
                return self.current_schedule
            if key == "next_schedule":
                return self.next_schedule
            if key == "replace_dict":
                return self.replace_dict
            if key == "ignore_subjects":
                return self.ignore_subjects
            if key == "today_schedule":
                return self.today_schedule
            return None
        else:
            return result

    @error_handler    
    def member_wxid(self, name:str, active: bool = True):
        """加载会员"""
        members = self.get_cache_data("members")
        if members is None:
            return ''
        if active:
            wxid = members[(members["alias"] == name) & (members["active"] == active)]["wxid"].tolist()
        else:
            wxid = members[(members["alias"] == name)]["wxid"].tolist()
        return '' if not wxid else wxid[0]

    @error_handler
    def get_wxids(self, name, active: bool = True):
        wxids = []
        class_template = self.get_cache_data("class_template")
        teacher_template = self.get_cache_data("teacher_template")
        if class_template is None or teacher_template is None:
            return wxids
        
        if active:
            teacher_template["active"] = teacher_template["active"].astype(int)
            teacher_template = teacher_template[teacher_template["active"] == 1]

        if name in class_template["class_name"].tolist():
            class_leaders = dict(zip(class_template["class_name"], class_template["leaders"]))
            leaders = class_leaders[name].split("/")
            for leader in leaders:
                wxids.append(self.member_wxid(leader, active=active, notify=True))
            return wxids
        elif name in teacher_template["name"].tolist():
            wxids.append(self.member_wxid(name, active=active, notify=True))
            return wxids
        else:
            return wxids

    def get_sid(self, cname:str, name:str):
        """获取学生ID"""
        students = self.get_cache_data("students")
        if students is None:
            return ''
        sid = students[(students["cname"] == cname) & (students["name"] == name)]["sid"].tolist()
        return '' if not sid else sid[0]

    def get_alias(self, wxid:str):
        """获取wxid对应的姓名"""
        members = self.get_cache_data("members")
        if members is None:
            return ''
        alias = members[members["wxid"] == wxid]["alias"].tolist()
        return '' if not alias else alias[0]

    @staticmethod
    @error_handler
    # 此方法集成到后面的 update_schedule 方法中
    def move_file(src, dst):
        """移动文件"""
        if not os.path.exists(src):
            log.error(f"源文件不存在: {src}")
            return 0
        os.makedirs(os.path.dirname(dst), exist_ok=True)
        shutil.move(src, dst)
        log.info(f"已移动文件: {src} -> {dst}")
        return 1

    def get_subject_teacher(self, subject: str) -> str:
        """获取学科对应的老师"""
        teacher_template = self.get_cache_data("teacher_template")
        if teacher_template is None:
            return subject
        subject_teacher = teacher_template.apply(list, axis=1)
        for s_t in subject_teacher:
            if subject in s_t[1].split("/"):
                return s_t[0]
        return subject
    
    def replace_subject_teacher(self, df_schedule: pd.DataFrame, teacher_flag: bool = True) -> pd.DataFrame:
        """
        将课表中的学科替换为对应的老师
        Args:
            df_schedule: 课表DataFrame
            teacher_flag: 是否替换为老师名字
        Returns:
            pd.DataFrame: 替换后的课表
        """
        teacher_template = self.get_cache_data("teacher_template")
        if teacher_template is None:
            return df_schedule
        subject_to_teacher = {}
        for _, row in teacher_template.iterrows():
            for subject in row["subject"].split("/"):
                if teacher_flag:
                    subject_to_teacher[subject] = row["name"]
                else:
                    subject_to_teacher[subject] = subject[:2]
        def replace_with_teacher(x):
            if not isinstance(x, str) or x == "-":
                return x
            return subject_to_teacher.get(x, x)

        return df_schedule.map(replace_with_teacher)
    
    def _title_to_week(self, title: str) -> list:
        """将课表标题转换为周历日期"""
        title_date = re.match(r"(\d+)",  title)
        if title_date is None or len(title_date.group(1)) != 8:
            return []
        title_date = title_date.group(1)
        week_info = self.get_cache_data("week_info")
        for week in week_info:
            if title_date == week[1]:
                return week
        return []
    
    def _check_schedule_date(self, df_schedule: pd.DataFrame, title:str) -> str:
        """检查课表日期是否符合要求"""
        if "date" not in df_schedule.columns:
            return "课表中没有日期列"
        week = self._title_to_week(title)
        if not week:
            return "课表文件名日期错误"
        monday = datetime.strptime(week[1], "%Y%m%d")
        date_list = []
        for i in range(7):
            date_list.append(int((monday + timedelta(days=i)).strftime("%d")))
        # df_date = df_schedule["date"].astype(int).tolist()
        df_date = pd.to_numeric(df_schedule["date"], errors="coerce").dropna().astype(int).tolist()
        df_date = list(dict.fromkeys(df_date))
        invalid_dates = [d for d in df_date if d not in date_list]
        if invalid_dates:
            return f"课表中日期 {invalid_dates} 不在当前周内"
        return "ok"

    def _check_schedule_class(self, df_schedule: pd.DataFrame) -> str:
        """检查课表班级是否符合要求"""
        class_template = self.get_cache_data("class_template")
        if class_template.empty:
            return "班级模板为空"
        class_names = class_template["class_name"].tolist()
        missing_classes = [c for c in class_names if c not in df_schedule.columns]
        if missing_classes:
            return f"课表中班级 {missing_classes} 不在班级模板中"
        return "ok"

    def _check_repeated_subjects(self, df_schedule: pd.DataFrame) -> str:
        """检查课表中是否有重复学科"""
        df_teacher = self.replace_subject_teacher(df_schedule)
        class_list = self.get_cache_data("class_template")["class_name"].tolist()
        df_class = df_teacher[class_list]
        repeated = self.get_cache_data("ignore_subjects")
        repeated_linses = []
        for index, row in df_class.iterrows():
            duplicates = row[row.duplicated()].to_dict()
            if duplicates:
                for k, v in duplicates.items():
                    if v != "-" and v:
                        subject = df_schedule.loc[index, k]
                        if subject not in repeated:
                            repeated_linses.append(f"第{index+2}行: {k}:{v}!")
        
        if repeated_linses:
            return f"课表中存在重复学科:" + "\n".join(repeated_linses)
        return "ok"
    
    # @error_handler
    def check_schedule(self, df_schedule: pd.DataFrame, title:str) -> str:
        """检查课表是否符合要求, 请先将df_schedule格式化"""
        if df_schedule is None or df_schedule.empty:
            return "读取课表失败"
        result = self._check_schedule_date(df_schedule, title)
        if result != "ok":
            return result
        result = self._check_schedule_class(df_schedule)
        if result != "ok":
            return result
        result = self._check_repeated_subjects(df_schedule)
        if result != "ok":
            return result
        return "ok"
    
    # 此方法集成到后面的 update_schedule 方法中
    def update_schedule_file(self, schedule_file:str, title:str, new_name:str):
        """课表检查无误后, 更新课表文件"""
    
        c_month = self.get_cache_data("c_month")
        schedule_dir = os.path.join(self.lesson_dir, c_month, "class_schedule")
        week = self._title_to_week(title)
        if not os.path.exists(schedule_dir):
            self.create_c_month_dir()
        schedule_list = self.get_cache_data("schedule_file")
        old_file = schedule_list[1] if week[4] else schedule_list[0]
        new_file = os.path.join(schedule_dir, new_name)
        new_file = new_file.replace("课表", "schedule")
        # schedule_file 复制到 class_schedule 目录
        try:
            shutil.copy(schedule_file, new_file)
            log.info(f"已更新课表文件: {new_file}")
            # 旧文件 移动到 history 目录
            dst = old_file.replace("class_schedule", "schedule_history")
            # print(old_file, dst)
            if old_file:
                if self.move_file(old_file, dst):
                    self.notify_admins(f"更新课表成功: {new_file}", log_level="info")
                    return new_file
                else:
                    self.notify_admins(f"更新课表失败: 移动旧课表文件到历史文件夹失败 {old_file}", log_level="error")
                    return None
            else:
                return new_file
        except Exception as e:
            self.notify_admins(f"更新课表失败: {e}", log_level="error")
            return None
    
    def schedule_diff(self, old_df, new_df):
        """比较课表差异"""
        old_df_teacher = self.replace_subject_teacher(old_df)
        new_df_teacher = self.replace_subject_teacher(new_df)
        class_list = self.get_cache_data("class_template")["class_name"].tolist()
        # 创建一个与new_df相同形状的DataFrame，用于存储差异
        diff_df = pd.DataFrame(index=new_df.index, columns=new_df.columns)

        # 比较old_df 和 new_df, 标记差异
        for col in class_list:
            for idx in new_df_teacher.index:
                if old_df_teacher.loc[idx, col] != new_df_teacher.loc[idx, col]:
                    diff_df.loc[idx, col] = (f"{old_df_teacher.loc[idx, col]} -> {new_df_teacher.loc[idx, col]}")
        # 将diff_df 转换为字典
        diff_dict = diff_df.to_dict()
        changes = []
        for col, row_dict in diff_dict.items():
            for idx, value in row_dict.items():
                if "->" in str(value):
                    old, new = value.split(" -> ")
                    date = new_df.loc[idx, "date"]
                    week = new_df.loc[idx, "week"]
                    order = new_df.loc[idx, "order"]
                    changes.append([idx, col, date, order, old, new, week])
        
        # 将变更按照 班级和老师 进行分类
        class_changes = []
        for change in changes:
            category = change[1]
            if category not in class_changes:
                class_changes.append(category)

        diff_teachers = []
        for change in changes:
            old_category = change[4]
            new_category = change[5]
            if old_category not in diff_teachers:
                teacher = self.get_subject_teacher(old_category)
                diff_teachers.append(teacher)
            if new_category not in diff_teachers:
                teacher = self.get_subject_teacher(new_category)
                diff_teachers.append(teacher)
        return class_changes, diff_teachers
    
    @error_handler
    def get_class_schedule(self, class_name:str, week_next:bool=False) -> pd.DataFrame:
        """获取班级课表"""
        if week_next:
            schedule_data = self.get_cache_data("next_schedule")
        else:
            schedule_data = self.get_cache_data("current_schedule")
        if schedule_data is None:
            return pd.DataFrame()
        required_columns = ["date", "week", "order", class_name]
        df = self.replace_subject_teacher(schedule_data, teacher_flag=False)
        df = df[required_columns]
        grouped_df = df.groupby("week")[class_name].apply(list).reset_index()
        new_df = pd.DataFrame(grouped_df[class_name].tolist(), index=grouped_df["week"])
        new_df.columns = df["order"][: new_df.shape[1]]
        new_df.index.name = "星期"
        new_df.columns.name = "节次"
        new_df = new_df.map(lambda x: "-" if x is None else x)
        return new_df.T
    
    def get_teacher_schedule(self, teacher_name:str, week_next:bool=False) -> pd.DataFrame:
        """获取教师课表"""
        if week_next:
            schedule_data = self.get_cache_data("next_schedule")
        else:
            schedule_data = self.get_cache_data("current_schedule")
        if schedule_data is None:
            return None
        df_subject = self.replace_subject_teacher(schedule_data, teacher_flag=False)
        df = self.replace_subject_teacher(schedule_data, teacher_flag=True)
        grouped_df = df[["week", "order"]].groupby("week")
        max_l_group = max(grouped_df, key=lambda x: len(x[1]))
        max_l_week = max_l_group[0]

        schedule_order = df[df["week"] == max_l_week]["order"].tolist()
        week_list = list(grouped_df.groups.keys())
        teacher_df = pd.DataFrame(columns=week_list, index=schedule_order)
        teacher_schedule = df[df.isin([teacher_name]).any(axis=1)]
        for index, row in teacher_schedule.iterrows():
            week = row["week"]
            order = row["order"]
            teacher_column = row.index[row == teacher_name][0]
            teacher_subject = df_subject.loc[index, teacher_column]
            teacher_df.loc[order, week] = (f"{teacher_column}-{teacher_subject}" if teacher_subject else teacher_column)
        return teacher_df

    def df_to_png(self, df: pd.DataFrame, png_name:str="temp.png", title:str="", index_name:str="节次\星期"):
        """
        将 DataFrame 转换为 PNG 图片
        
        参数:
            df: 要转换的 DataFrame
            png_name: 输出的 PNG 文件名
            title: 图片标题
            index_name: 索引列的名称
            
        返回:
            包含保存路径的列表
        """
        from matplotlib.figure import Figure
        from matplotlib.backends.backend_agg import FigureCanvasAgg
        
        # 复制 DataFrame 避免修改原始数据，并将 NaN 替换为空字符串
        df = df.copy()
        df = df.fillna('')
        df.index.name = index_name
        df.reset_index(inplace=True)
        
        # 生成带时间戳的唯一文件名
        timestamp = str(int(time.time() * 1000))
        g = png_name.split(".")
        png_name = f"{g[0]}_{timestamp}.{g[1]}"
        
        save_path = os.path.join(self.lesson_dir, "temp", png_name)
        
        # 获取表格行列数，用于计算画布大小
        n_rows = len(df)
        n_cols = len(df.columns)
        
        # 创建画布（使用面向对象 API，避免 pyplot 全局状态）
        fig = Figure(figsize=(min(n_cols*1.1, 18), n_rows * 0.18 + 1))
        canvas = FigureCanvasAgg(fig)
        ax = fig.add_subplot(111)
        ax.axis('off')  # 隐藏坐标轴
        
        # 添加标题（如果有）
        if title:
            fig.suptitle(title, fontsize=12, fontweight='bold', y=0.95, color='#000000')
        
        # 创建表格
        table = ax.table(
            cellText=df.values,        # 单元格数据
            colLabels=df.columns,       # 列标题
            loc='center',               # 表格位置居中
            cellLoc='center'            # 单元格内容居中
        )
        
        # 设置字体大小和表格缩放比例
        table.auto_set_font_size(False)
        table.set_fontsize(10)
        table.scale(0.9, 1.1)  # 宽度缩放 0.9，高度缩放 1.1
        
        # 遍历所有单元格，设置样式
        for key, cell in table.get_celld().items():
            row, col = key
            cell.set_edgecolor('#000000')  # 边框颜色：黑色
            cell.set_linewidth(0.5)          # 边框宽度：1px
            cell.set_text_props(fontfamily=get_cached_chinese_font())  # 中文字体
            
            if row == 0:
                # 表头样式：蓝色背景，白色粗体文字
                cell.set_facecolor('#4A90E2')
                cell.set_text_props(color='white', fontweight='bold', fontsize=10)
                cell.set_edgecolor('#2E5C8A')  # 表头边框深色
            elif row == 1:
                cell.set_facecolor('#FFFFFF') 
            else:
                # 数据行：每 4 行一组，组内颜色相同，组间交替（浅蓝色和白色）
                group = (row - 2) // 4  # 计算属于第几组（从 0 开始）
                if group % 2 == 0:
                    cell.set_facecolor('#E8F4FF')  # 偶数组：浅蓝色
                else:
                    cell.set_facecolor('#FFFFFF')  # 奇数组：白色
                
                cell.set_text_props(color='#000000', fontsize=10)
        
        # 调整表头高度
        for col_idx in range(len(df.columns)):
            cell = table[(0, col_idx)]
            cell.set_height(0.1)
        
        # 保存为 PNG 图片
        fig.tight_layout()
        canvas.print_figure(save_path, dpi=200, bbox_inches='tight', facecolor='white', edgecolor='none')
        
        return [save_path]
        
    
    def current_teaching(self) -> dict:
        """获取当前正在上的课程"""
        df = self.get_cache_data("today_schedule")
        if df is None or df.empty:
            return {}
        current_time = datetime.now().strftime("%H:%M")
        current_classes = {}
        current_period = None
        time_tables = self.get_cache_data("time_table")
        periods = {order: show_time for order, show_time in zip(time_tables["order"], time_tables["show_time"])}
        for period, time_range in periods.items():
            start_time, end_time = time_range.split("-")
            start_minutes = sum(int(x)*60**i for i ,x in enumerate(reversed(start_time.split(":"))))
            end_minutes = sum(int(x)*60**i for i ,x in enumerate(reversed(end_time.split(":"))))
            current_minutes = sum(int(x)*60**i for i ,x in enumerate(reversed(current_time.split(":"))))
            if start_minutes <= current_minutes <= end_minutes:
                current_period = period
                break
        class_list = self.get_cache_data("class_template")["class_name"].tolist()
        if current_period is not None:
            # 确保类型一致，统一转换为字符串比较
            df["order"] = df["order"].astype(str)
            df_current = df[df["order"] == str(current_period)]
            if df_current.empty:
                return current_classes
            for cls_name in class_list:
                if cls_name in df_current.columns:
                    values = df_current[cls_name].values
                    if len(values) > 0:
                        current_classes[cls_name] = values[0]
        return current_classes

def clear_temp_file(folder="temp"):
    l = Lesson()
    temp_dir = os.path.join(l.lesson_dir, folder)
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir)
        log.info(f"已删除临时文件夹: {temp_dir}")
        time.sleep(1)
        os.makedirs(temp_dir, exist_ok=True)
        log.info(f"已创建临时文件夹: {temp_dir}")
        l.notify_admins(f"已清空临时文件夹: {temp_dir}")
    else:
        os.makedirs(temp_dir, exist_ok=True)
        log.info(f"临时文件夹不存在,已创建: {temp_dir}")
        l.notify_admins(f"临时文件夹不存在,已创建: {temp_dir}")

async def create_month_dir():
    """创建当前月份的文件夹, 并将上课月的课表复制到 新的月份 目录下"""
    l = Lesson()
    current_dir = l.create_c_month_dir()
    if current_dir:
        log.info(f"已创建当前月份文件夹: {current_dir}")
    else:
        l.notify_admins(f"创建当前月份文件夹失败", log_level="error")
        return False
    c_month = current_dir.split("/")[-1].split("\\")[-1]
    year = int(c_month[:4])
    month = int(c_month[4:])
    if month == 1:
        p_month = f"{year-1}12"
    else:
        p_month = f"{year}{month-1:02d}"
    p_month_dir = os.path.join(l.lesson_dir, p_month)
    c_month_dir = os.path.join(l.lesson_dir, c_month)
    schedule_list = os.listdir(os.path.join(p_month_dir, "class_schedule"))
    time_flag = 0
    schedule_file = ""
    for schedule in schedule_list:
        timestamp = int(schedule.split("-")[-1].split(".")[0])
        if timestamp > time_flag:
            time_flag = timestamp
            schedule_file = schedule
    if schedule_file:
        p_schedule_file = os.path.join(p_month_dir, "class_schedule", schedule_file)
        c_schedule_file = os.path.join(c_month_dir, "class_schedule", schedule_file)
        shutil.copy(p_schedule_file, c_schedule_file)
        log.info(f"已复制上月份课表到当前月份文件夹: {c_schedule_file}")
        if l.refresh_cache():
            return True
    else:
        l.notify_admins(f"上月份课表不存在", log_level="error")
        return False

async def process_and_send_image(lesson: Lesson, df: pd.DataFrame, png_name: str, title: str, wxid: str, producer: str, tips=False):
    """异步处理并发送图片"""
    with ThreadPoolExecutor() as executor:
        df_png = await asyncio.get_event_loop().run_in_executor(executor, lesson.df_to_png, df, png_name, title)
        if df_png:
            pic_path = df_png[0][len(lesson.lesson_dir):].replace("\\", "/")
            if tips:
                send_text(f"你的课有调整，请注意查看！", wxid)
            send_image(pic_path, wxid, producer)
            return True
        else:
            lesson.notify_admins(f"生成课表图片失败: {title}", log_level="error")
            return False

async def process_class_schedule(lesson: Lesson, df: pd.DataFrame, png_name: str, title: str, class_name: str, producer: str, tips=False):
    """异步处理并发送班级课表图片"""
    with ThreadPoolExecutor() as executor:
        class_pic = await asyncio.get_event_loop().run_in_executor(executor, lesson.df_to_png, df, png_name, title)
        if class_pic:
            pic_path = class_pic[0][len(lesson.lesson_dir):].replace("\\", "/")
            wxids = lesson.get_wxids(class_name)
            for wxid in wxids:
                if tips:
                    send_text(f"{class_name}的课有调整，请注意查看！", wxid)
                send_image(pic_path, wxid, producer)
            return True
        else:
            lesson.notify_admins(f"生成班级课表图片失败: {title}", log_level="error")
            return False

async def _update_schedule(l: Lesson, title: str, temp_file: str, new_name: str, content=''):
    # 读取课表文件, 检查课表是否符合要求
    week = l._title_to_week(title)
    if not week:
        l.notify_admins(f"解析课表标题失败: {title}", log_level="error")
        return False
    week_flag = "双" if week[0]%2 == 0 else "单"
    schedule_data = l.load_excel_file(temp_file)
    schedule_data = l.format_schedule(schedule_data, week_flag, ignore=True)
    rsp = l.check_schedule(schedule_data, title)
    if rsp != "ok":
        l.notify_admins(f"{rsp}", log_level="error")
        return False
    # 更新课表文件
    rsp = l.update_schedule_file(temp_file, title, new_name)
    # print(rsp)
    if not rsp:
        return False
    diff_flag = 1 if "微调" in content else 0
    if diff_flag:
        schedule = "next_schedule" if week[4] else "current_schedule"
        old_schedule = l.get_cache_data(schedule)
        if not l.refresh_cache():
            return False
        new_schedule = l.get_cache_data(schedule)
        diffs = l.schedule_diff(old_schedule, new_schedule)
        if diffs != ([], []):
            week_next = True if week[4] else False
            week_flag = "下周" if week_next else "本周"
            teachers = []
            errors = []
            task = []
            class_diff = diffs[0]
            teachers_dff = diffs[1]
            for k in class_diff:
                class_df = l.get_class_schedule(k, week_next)
                title = f"{k}{week_flag}的课表"
                teachers.append(k)
                if not class_df.empty:
                    task.append(process_class_schedule(l, class_df, f"{k}.png", title, k, "lesson", True))
            for k in teachers_dff:
                teacher_df = l.get_teacher_schedule(k, week_next)
                title = f"{k}{week_flag}的课表"
                wxids = l.get_wxids(k)
                if not wxids or wxids ==[""]:
                    errors.append(k)
                else:
                    teachers.append(k)
                    for wxid in wxids:
                        task.append(process_and_send_image(l, teacher_df, f"{wxid}.png", title, wxid, "lesson", True))
            if task:
                await asyncio.gather(*task)
            teachers = set(teachers)
            tips = f"{week_flag}课表微调，已通知以下老师："
            for t in teachers:
                tips += f"\n{t} "
            l.notify_admins(tips, log_level="info")
            if errors:
                tips = "以下老师未通知："
                for e in errors:
                    tips += f"\n{e} "
                l.notify_admins(tips, log_level="error")
            return True
        else:
            l.notify_admins(f"课表未发生变化\n{title}", log_level="info")
            return True
    else:
        if not l.refresh_cache():
            l.notify_admins(f"刷新课表缓存失败\n{title}", log_level="error")
            return False
        l.notify_admins(f"{title}已上传成功，发送如下指令通知老师")
        if week[4]:
            l.notify_admins(f"更新下周的课表", log_level="info")
        else:
            l.notify_admins(f"更新本周的课表", log_level="info")
        return True


async def update_schedule(record: any):
    # title = "[文件] 课表20251020微调.xlsx"
    msg_id = record.msg_id
    content = record.content
    title = re.match(r"^\[文件\] 课表(\d+).*\.xlsx$", content).group(1)
    new_name = f"课表{title}-{int(time.time())}.xlsx"
    l = Lesson()
    temp_file = os.path.join(l.lesson_dir, "temp", new_name)
    # 下载课表文件 到 临时文件夹
    response = down_file(msg_id, temp_file)
    if not response:
        l.notify_admins(f"下载课表失败: {title}", log_level="error")
        return False
    # 更新课表
    rsp = await _update_schedule(l, title, temp_file, new_name, content)
    if not rsp:
        l.notify_admins(f"更新课表失败: {title}", log_level="error")
        return False
    return True

async def update_schedule_all(record: any):
    """
    更新所有班级和老师的课表
    """
    l = Lesson()
    content = record.content
    s_name = re.match(r"^(.+?)\s*的课表$", content).group(1)
    if s_name == "更新本周":
        df = l.get_cache_data("current_schedule")
        if df.empty:
            l.notify_admins(f"当前周的课表为空", log_level="error")
            return False
        week_next = False
    elif s_name == "更新下周":
        df = l.get_cache_data("next_schedule")
        if df.empty:
            l.notify_admins(f"下周的课表为空", log_level="error")
            return False
        week_next = True
    else:
        l.notify_admins(f"未知的课表更新请求：{content}", log_level="error")
        return False
    teacher_template = l.get_cache_data("teacher_template")
    class_template = l.get_cache_data("class_template")
    teachers = teacher_template[teacher_template["active"] == True]["name"].tolist()
    classes = class_template[class_template["active"] == True]["class_name"].tolist()
    tasks = []
    for t in teachers:
        wxids = l.get_wxids(t)
        if not wxids:
            l.notify_admins(f"未找到老师 {t} 的微信ID", log_level="error")
            continue
        wxid = wxids[0]
        teacher_df = l.get_teacher_schedule(t, week_next)
        if teacher_df.empty:
            l.notify_admins(f"老师 {t} 的课表为空", log_level="info")
            continue
        title = f"{t}{'下周' if week_next else '本周'}的课表"
        tasks.append(process_and_send_image(l, teacher_df, f"{wxid}.png", title, wxid, "lesson"))
    
    for c in classes:
        class_df = l.get_class_schedule(c, week_next)
        if class_df.empty:
            l.notify_admins(f"班级 {c} 的课表为空", log_level="info")
            continue
        title = f"{c}{'下周' if week_next else '本周'}的课表"
        tasks.append(process_class_schedule(l, class_df, f"{c}.png", title, c, "lesson"))
    
    if tasks:
        await asyncio.gather(*tasks)

@check_permission
async def teacher_schedule(record: any):
    """获取指定老师或班级的课表"""
    content = record.content
    if len(content) > 9:
        return False
    wxid = record.roomid
    l = Lesson()
    name = re.match(r"^(.*?)\s*的课表$", content).group(1)
    week_next = True if "下周" in name else False
    name = name.replace("下周", "")
    if name[0]=="高" and name[-1]=="班":
        class_df = l.get_class_schedule(name, week_next)
        if class_df.empty:
            send_text(f"班级 {name} 的课表为空", wxid, "", "lesson")
            return False
        df_png = l.df_to_png(class_df, f"{wxid}.png", title=f"{name}{'下周' if week_next else '本周'}的课表")[0]
        pic_path = df_png[len(l.lesson_dir):].replace("\\", "/")
        send_image(pic_path, wxid, "lesson")
        return True
    if name == "我":
        members = l.members[l.members["active"] == True]
        member_dict = dict(zip(members["alias"], members["wxid"]))
        for k, v in member_dict.items():
            if wxid == v:
                name = k
                break
    df = l.get_teacher_schedule(name, week_next)
    if df.empty:
        send_text(f"您的课表为空", wxid, "", "lesson")
        return False
    df_png = l.df_to_png(df, f"{wxid}.png", title=f"{name}{'下周' if week_next else '本周'}的课表")[0]
    pic_path = df_png[len(l.lesson_dir):].replace("\\", "/")
    send_image(pic_path, wxid, "lesson")
    return True

@check_permission
async def get_current_schedule(record: any):
    """获取当前本周或下周的课表文件"""
    content = record.content
    week_next = 1 if "下周" in content else 0
    l = Lesson()
    schedule_file = l.get_cache_data("schedule_file")[week_next]
    if not schedule_file:
        send_text(f"未找到{'下周' if week_next else '本周'}的课表文件", record.roomid, "", "lesson")
        return True
    pre_path = l.lesson_dir
    schedule_file = schedule_file[len(pre_path):].replace("\\", "/")
    send_file(schedule_file, record.roomid, "lesson")
    static_url = config.get_config("static_url", "wechat.yaml")
    send_text(f"{static_url}{schedule_file}", record.roomid, "", "lesson")
    return True

@error_handler
def send_today_schedule(wxid="", notify=True, target_date=None):
    """发送今日课表（支持传入日期参数）"""
    l = Lesson()
    if not wxid:
        wxid = l.admin[0]
    
    if target_date is None:
        now = l.get_cache_data("now")
        target_date = now.strftime("%Y%m%d")
    
    if isinstance(target_date, str) and len(target_date) == 8:
        today = target_date
    else:
        try:
            if isinstance(target_date, str):
                target_dt = datetime.strptime(target_date, "%Y%m%d")
            elif isinstance(target_date, int):
                # 假设传入的是周几（1-5）
                now = l.get_cache_data("now")
                target_dt = now + timedelta(days=(target_date - now.weekday() - 1) % 7)
            else:
                target_dt = target_date
            today = target_dt.strftime("%Y%m%d")
        except:
            now = l.get_cache_data("now")
            today = now.strftime("%Y%m%d")
    
    today_df = l.get_weekday_schedule(target_dt if isinstance(target_date, int) else target_date)
    
    if today_df.empty:
        send_text(f"{today} 课表为空", wxid, "", "lesson")
        return True
    today_df.drop(["style", "date", "week"], axis=1, inplace=True, errors="ignore")
    today_df = today_df.set_index("order")
    def take_first_four(x):
        if isinstance(x, str) and len(x) >= 4:
            return x[:4]
        return x
    today_df = today_df.map(take_first_four)
    df_png = l.df_to_png(today_df, f"{wxid}.png", title=f"{today}的课表")[0]
    pic_path = df_png[len(l.lesson_dir):].replace("\\", "/")
    send_image(pic_path, wxid, "lesson")
    return True

@check_permission
async def get_today_schedule(record: any):
    """获取今日课表或指定周几的课表"""
    wxid = record.roomid
    content = record.content
    
    weekday_map = {"周一": 1, "周二": 2, "周三": 3, "周四": 4, "周五": 5}
    
    target_weekday = None
    for day_name, day_num in weekday_map.items():
        if day_name in content:
            target_weekday = day_num
            break
    
    send_today_schedule(wxid, target_date=target_weekday)
    return True

def refresh_cache_data(notify=False):
    """刷新cache数据"""
    l = Lesson()
    r = l.refresh_cache()
    if r:
        send_today_schedule(notify=notify)
        log.info(f"刷新课表数据成功")
    else:
        log.error(f"刷新课表数据失败")
    return r

async def refresh_schedule(record=None):
    """刷新cache数据"""
    if record:
        print("手动刷新")
        notify = True
    else:
        notify = False
    r = refresh_cache_data(notify=notify)
    return r

async def get_current_teacher(record: any):
    """获取当前正在上课教师"""
    l = Lesson()
    wxid = record.roomid
    teachers = l.current_teaching()
    if teachers == {}:
        send_text(f"暂无正在上课教师", wxid, "", "lesson")
        return True
    else:
        tips = "当前正在上课的教师有："
        for k, v in teachers.items():
            tips += f"\n{k}: {v}"
        send_text(tips, wxid, "", "lesson")
        return True

def today_teachers():
    now = datetime.now().strftime("%Y%m%d")
    l = Lesson()
    schedule_now = l.get_cache_data("now").strftime("%Y%m%d")
    log.info(f"schedule_now: {schedule_now}, now: {now}")
    if schedule_now != now:
        l.notify_admins(f"今日课表尚未更新，请检查配置", log_level="error")
        return
    today_df = l.get_cache_data("today_schedule")
    if today_df.empty:
        l.notify_admins(f"今日无课！", log_level="info")
        return
    time_table = l.get_cache_data("time_table")
    class_order = time_table["order"].tolist()
    today_df["order"] = today_df["order"].astype(str)
    teachers = {}
    for order in class_order:
        for class_name in l.get_cache_data("class_template")["class_name"].tolist():
            class_df = today_df[[class_name, "order"]]
            try:
                subjcet = class_df[class_df["order"] == str(order)][class_name].values[0]
                teacher = l.get_subject_teacher(subjcet)
                order_label = time_table[time_table["order"] == order]["label"].values[0]
                if teacher not in teachers:
                    teachers[teacher] = []
                teachers[teacher].append(f"{order_label}:{class_name} {subjcet[:2]}")
            except Exception as e:
                print(e, order, class_name)
    if not teachers:
        l.notify_admins(f"今日无课！", log_level="info")
        return
    errors = []
    for k, v in teachers.items():
        wxids = l.get_wxids(k)
        if wxids:
            tips = f"您今天有{len(v)}节课："
            for item in v:
                tips += f"\n{item}"
            send_text(tips, wxids[0], "", "lesson")
        else:
            if k not in "早读, -, 专业课, 英语, 活动, 自习, 校本":
                errors.append(k)
    if errors:
        l.notify_admins(f"以下教师没有微信ID：{', '.join(errors)}", log_level="warning")
    else:
        l.notify_admins(f"今日上课老师已经通知", log_level="info")

@check_permission
async def current_week_info(record: any):
    l = Lesson()
    send_text(str(l.week_info[0]), record.roomid, "", "lesson")

def group_send(xlsx_file, sender):
    df = pd.read_excel(xlsx_file, engine="openpyxl")
    cnt = 0
    l = Lesson()
    failed = []
    try:
        for _, row in df.iterrows():
            name = row["接收人"]
            wxids = l.get_wxids(name, notify=False)
            for wxid in wxids:
                if row["类型"] == "消息":
                    send_text(row["消息内容"], wxid, "", "lesson")
                elif row["类型"] == "课表":
                    week_next = True if "下周" in row["消息内容"] else False
                    if name[0] == "高" and name[-1] == "班":
                        class_df = l.get_class_schedule(name, week_next=week_next)
                        df_png = l.df_to_png(class_df, f"{wxid}.png", title=f"{name}{'下周' if week_next else '本周'}的课表")[0]
                        if not df_png:
                            log.error(f"生成{name}{'下周' if week_next else '本周'}的课表失败")
                            failed.append(name)
                            continue
                        pic_path = df_png[len(l.lesson_dir):].replace("\\", "/")
                        send_image(pic_path, wxid, "lesson")
                    else:
                        df = l.get_teacher_schedule(name, week_next=week_next)
                        if df.empty:
                            log.error(f"{name}{'下周' if week_next else '本周'}的课表为空")
                            failed.append(name)
                            continue
                        df_png = l.df_to_png(df, f"{wxid}.png", title=f"{name}{'下周' if week_next else '本周'}的课表")[0]
                        if not df_png:
                            log.error(f"生成{name}{'下周' if week_next else '本周'}的课表失败")
                            failed.append(name)
                            continue
                        pic_path = df_png[len(l.lesson_dir):].replace("\\", "/")
                        send_image(pic_path, wxid, "lesson")
            cnt += 1
    except KeyError as e:
        log.error(f"KeyError：{str(e)}")
    tips = f"发送{cnt}条消息，{len(failed)}条失败"
    if failed:
        tips += f"\n未通知以下人员："
        for f in failed:
            tips += f"\n{f}"
    else:
        tips += "\n所有人员都已通知"
    send_text(tips, sender, "", "lesson")

async def mass_message(record: any):
    """群发消息"""
    l = Lesson()
    content = record.content
    title = re.match(r"^\[文件\] ((学发|教发)群发通知\d*\.xlsx)$", content).group(1)
    if title:
        title = title.split(".")[0]
        title = re.sub(r"\d+", "", title)
        notify_file = f"{title}-{time.strftime('%Y%m%d%H%M%S', time.localtime())}.xlsx"
        new_file = os.path.join(l.lesson_dir, "notice", notify_file)
        rsp_path = down_file(record.msg_id, new_file)
        if rsp_path == "":
            send_text("通知文件下载失败，请重新发送！", record.roomid)
            return
        else:
            group_send(rsp_path, record.roomid)

@check_permission
async def file_template(record: any):
    """获取文件， 根据 file_template 文件配置 文件字典"""
    file_name = (record.content.replace("：", ":").replace(" ", "").replace("获取文件:", "").replace("文件获取:", ""))
    file_template_path = config.get_config("file_template", "message.yaml")[file_name]
    template_file = "template/" + file_template_path
    if template_file.split(".")[-1] in ["jpg", "png", "jpeg", "gif"]:
        send_image(template_file, record.roomid)
    else:
        send_file(template_file, record.roomid)

async def pan_share(record: any):
    share_links = config.get_config("pan_share", "message.yaml")
    if share_links:
        links = "当前分享链接如下："
        for link in share_links:
            links += f"\n{link}"
        send_text(links, record.roomid)

@check_permission
async def sunday_record(record: any):
    app_xml = config.get_config("周日值班记录", "file_template.yaml")
    if not app_xml:
        send_text("请先配置 值日值班记录 模板", record.roomid, "", "lesson")
        return 0
    content = record.content
    receives = content.split("-")[1].split("+")
    l = Lesson()
    for r in receives:
        try:
            if r == "群":
                wxid = config.get_config("teach_chatroom", "lesson.yaml")
            else:
                wxid = l.get_wxids(r)[0]
            send_app_msg(app_xml, wxid, producer="lesson")
            time.sleep(3)
        except:
            send_text(f"发送{r}的值班记录失败", record.roomid, "", "lesson")
    return 1

async def schedule_tips(record: any):
    text = record.content
    if len(text) > 10:
        return
    tips = "抱歉，您输入的指令有误，请看一下这个文档，重新输入指令。\n指令指南：https://mp.weixin.qq.com/s/n6DkDvUeCzcnKIYZB7paOA"
    send_text(tips, record.roomid, "", "lesson")

def add_teacher(wxid: str, name: str, subject: str):
    """添加教师"""
    with Member() as m:
        result = m.member_info(wxid)
        if not result:
            try:
                m.insert_member(wxid, wxid, name, 1, 50, 0, 10, "basic/lesson")
                l = Lesson()
                file_path = os.path.join(l.lesson_dir, "checkTemplate.xlsx")
                wb = load_workbook(file_path)
                ws = wb["teachers"]
                ws.append([name, subject, subject[:2], 1, "666666"])
                wb.save(file_path)
                db_path = config.get_config("file_db")
                conn = sqlite3.connect(db_path)
                cursor = conn.cursor()
                pwd = hashlib.sha256(("FUSALT:666666").encode()).hexdigest()
                cursor.execute("INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)", (name, pwd, "user"))
                conn.commit()
                conn.close()
                l.refresh_cache()
                return "OK"
            except Exception as e:
                return str(e)
        else:
            return "该教师已存在"

def delete_teacher(name: str):
    """删除教师"""
    with Member() as m:
        result = m.member_wxid(name)
        if result:
            try:
                m.delte_member(result[0])
                l = Lesson()
                file_path = os.path.join(l.lesson_dir, "checkTemplate.xlsx")
                wb = load_workbook(file_path)
                ws = wb["teachers"]
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
                    if row[0].value == name:
                        ws.delete_rows(row[0].row)
                        break
                wb.save(file_path)
                db_path = config.get_config("file_db")
                conn = sqlite3.connect(db_path)
                cursor = conn.cursor()
                cursor.execute("DELETE FROM users WHERE username = ?", (name,))
                conn.commit()
                conn.close()
                return "OK"
            except Exception as e:
                return str(e)
        else:
            return "该教师不存在"

async def add_teacher_async(record):
    """异步添加教师"""
    contents = record.content.replace("：", ":").splitlines()
    if len(contents) != 3:
        send_text("请输入正确的教师信息，格式为：95+\n姓名:\n学科:", record.roomid, "", "lesson")
        return
    name = contents[1].strip()
    if not name or "姓名" not in name:
        send_text("请输入正确的教师信息，格式为：95+\n姓名:\n学科:", record.roomid, "", "lesson")
        return
    name = name.split(":")[-1]
    with Member() as m:
        wxid = m.remark_wxid("天龙"+name)
    if not wxid:
        send_text(f"{name}尚未添加小助手", record.roomid, "", "lesson")
        return
    subject = contents[2].strip()
    if not subject or "学科" not in subject:
        send_text("请输入正确的教师信息，格式为：95+\n姓名:\n学科:", record.roomid, "", "lesson")
        return
    subject = subject.split(":")[-1]
    result = add_teacher(wxid, name, subject)
    if result == "OK":
        send_text(f"添加教师 {name} 成功", record.roomid, "", "lesson")
    else:
        send_text(f"添加教师 {name} 失败：{result}", record.roomid, "", "lesson")

async def delete_teacher_async(record):
    """异步删除教师"""
    name = record.content.splitlines()[1].strip().replace("：", ":").split(":")[-1]
    result = delete_teacher(name)
    if result == "OK":
        send_text(f"删除教师 {name} 成功", record.roomid, "", "lesson")
    else:
        send_text(f"删除教师 {name} 失败：{result}", record.roomid, "", "lesson")

def lesson_pwd(record):
    # send_text = print
    l = Lesson()
    teachers = l.get_cache_data("teacher_template")
    # print(teachers)
    if teachers.empty:
        send_text("教师模板不存在", record.roomid, "", "lesson")
        return
    name = l.get_alias(record.sender)
    # 优先使用 raw_pwd (明文密码)，如果不存在则使用 pwd
    if 'raw_pwd' in teachers.columns:
        pwd_list = teachers[teachers['name'] == name]['raw_pwd'].values
    else:
        pwd_list = teachers[teachers['name'] == name]['pwd'].values
    if len(pwd_list) > 0:
        pwd = pwd_list[0]
        if pwd:
            send_text(f"{name}的密码为：{pwd}", record.sender, "", "lesson")
        else:
            send_text(f"{name}的密码不存在", record.sender, "", "lesson")
    else:
        send_text(f"未找到教师 {name}", record.sender, "", "lesson")

async def manual_update_schedule():
    """手动更新课程表"""
    # async def _update_schedule(l: Lesson, title: str, temp_file: str, new_name: str, content=''):
    l = Lesson()
    up_file = os.listdir(os.path.join(l.lesson_dir, l.c_month, "temp"))
    if ".DS_Store" in up_file:
        up_file.remove(".DS_Store")
    print(up_file)
    if not up_file or len(up_file) > 1:
        l.notify_admins("没有上传的课表文件或上传了多个文件", log_level="error")
        for f in up_file:
            os.remove(os.path.join(l.lesson_dir, l.c_month, "temp", f))
        return False
    # Filter for files starting with "课表" and ending with ".xlsx"
    schedule_files = [f for f in up_file if f.startswith("课表") and f.endswith(".xlsx")]
    if not schedule_files:
        l.notify_admins("没有符合格式的课表文件 (需以'课表'开头并以'.xlsx'结尾)", log_level="error")
        os.remove(os.path.join(l.lesson_dir, l.c_month, "temp", up_file[0]))
        return False
    
    # Sort to get the most recent one if multiple exist, or just take the first
    file_to_process = schedule_files[0]
    
    match = re.match(r"^课表(\d+).*\.xlsx$", file_to_process)
    if not match:
         l.notify_admins(f"文件名格式不符合预期: {file_to_process}", log_level="error")
         return False
    
    title = match.group(1)
    up_file_path = os.path.join(l.lesson_dir, l.c_month, "temp", file_to_process)
    new_name = f"课表{title}-{int(time.time())}.xlsx"
    temp_file = os.path.join(l.lesson_dir, l.c_month, "temp", new_name)
    shutil.move(up_file_path, temp_file)
    # 更新课表
    rsp = await _update_schedule(l, title, temp_file, new_name, file_to_process)
    # os.remove(up_file_path)  # 已被 shutil.move 移除，此处重复删除会导致 FileNotFoundError
    os.remove(temp_file)
    if not rsp:
        l.notify_admins(f"更新课表失败: {title}", log_level="info")
        return False
    return True
